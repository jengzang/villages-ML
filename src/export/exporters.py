"""
Export classes for different output formats.
"""

import csv
import json
import gzip
from abc import ABC, abstractmethod
from pathlib import Path
from typing import Any, Dict, List, Optional, Union
from datetime import datetime

from .formatters import format_number, format_percentage, format_timestamp, sanitize_latex


class BaseExporter(ABC):
    """Abstract base class for exporters."""

    def __init__(self):
        self.metadata: Dict[str, Any] = {}

    @abstractmethod
    def export(self, data: List[Dict[str, Any]], output_path: Union[str, Path],
               metadata: Optional[Dict[str, Any]] = None, **kwargs) -> None:
        """
        Export data to file.

        Args:
            data: List of dictionaries to export
            output_path: Output file path
            metadata: Optional metadata to embed
            **kwargs: Format-specific options
        """
        pass

    def _prepare_output_path(self, output_path: Union[str, Path]) -> Path:
        """Ensure output directory exists."""
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        return path

    def _merge_metadata(self, metadata: Optional[Dict[str, Any]]) -> Dict[str, Any]:
        """Merge provided metadata with defaults."""
        default_metadata = {
            'created_at': datetime.now().isoformat(),
            'exporter': self.__class__.__name__,
        }
        if metadata:
            default_metadata.update(metadata)
        return default_metadata


class CSVExporter(BaseExporter):
    """Export data to CSV format with UTF-8-BOM encoding."""

    def export(self, data: List[Dict[str, Any]], output_path: Union[str, Path],
               metadata: Optional[Dict[str, Any]] = None,
               compress: bool = False, **kwargs) -> None:
        """
        Export data to CSV file.

        Args:
            data: List of dictionaries to export
            output_path: Output file path
            metadata: Optional metadata to embed in header comments
            compress: Enable gzip compression
            **kwargs: Additional csv.writer options
        """
        if not data:
            raise ValueError("No data to export")

        path = self._prepare_output_path(output_path)
        merged_metadata = self._merge_metadata(metadata)

        # Add .gz extension if compressing
        if compress and not str(path).endswith('.gz'):
            path = Path(str(path) + '.gz')

        # Open file with appropriate handler
        open_func = gzip.open if compress else open
        mode = 'wt' if compress else 'w'

        with open_func(path, mode, encoding='utf-8-sig', newline='') as f:
            # Write metadata as comments
            for key, value in merged_metadata.items():
                f.write(f"# {key}: {value}\n")
            f.write(f"# Total rows: {len(data)}\n")
            f.write("#\n")

            # Write CSV data
            fieldnames = list(data[0].keys())
            writer = csv.DictWriter(f, fieldnames=fieldnames, **kwargs)
            writer.writeheader()
            writer.writerows(data)


class JSONExporter(BaseExporter):
    """Export data to JSON format with metadata."""

    def export(self, data: List[Dict[str, Any]], output_path: Union[str, Path],
               metadata: Optional[Dict[str, Any]] = None,
               pretty: bool = True, compress: bool = False, **kwargs) -> None:
        """
        Export data to JSON file.

        Args:
            data: List of dictionaries to export
            output_path: Output file path
            metadata: Optional metadata to embed
            pretty: Enable pretty-printing
            compress: Enable gzip compression
            **kwargs: Additional json.dump options
        """
        path = self._prepare_output_path(output_path)
        merged_metadata = self._merge_metadata(metadata)

        # Add .gz extension if compressing
        if compress and not str(path).endswith('.gz'):
            path = Path(str(path) + '.gz')

        # Prepare output structure
        output = {
            'metadata': merged_metadata,
            'data': data,
            'count': len(data)
        }

        # JSON dump options
        dump_kwargs = {
            'ensure_ascii': False,
            'indent': 2 if pretty else None,
            **kwargs
        }

        # Open file with appropriate handler
        open_func = gzip.open if compress else open
        mode = 'wt' if compress else 'w'

        with open_func(path, mode, encoding='utf-8') as f:
            json.dump(output, f, **dump_kwargs)


class ExcelExporter(BaseExporter):
    """Export data to Excel format with multiple sheets."""

    def export(self, data: List[Dict[str, Any]], output_path: Union[str, Path],
               metadata: Optional[Dict[str, Any]] = None,
               sheet_name: str = 'Data', **kwargs) -> None:
        """
        Export data to Excel file (single sheet).

        Args:
            data: List of dictionaries to export
            output_path: Output file path
            metadata: Optional metadata to embed
            sheet_name: Name of the worksheet
            **kwargs: Additional options
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        if not data:
            raise ValueError("No data to export")

        path = self._prepare_output_path(output_path)
        merged_metadata = self._merge_metadata(metadata)

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Write headers with formatting
        headers = list(data[0].keys())
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

        # Write data
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add metadata sheet
        meta_ws = wb.create_sheet('Metadata')
        meta_ws.append(['Key', 'Value'])
        for key, value in merged_metadata.items():
            meta_ws.append([key, str(value)])

        wb.save(path)

    def export_multi_sheet(self, output_path: Union[str, Path],
                          sheets: Dict[str, List[Dict[str, Any]]],
                          metadata: Optional[Dict[str, Any]] = None) -> None:
        """
        Export multiple datasets to different sheets.

        Args:
            output_path: Output file path
            sheets: Dictionary mapping sheet names to data
            metadata: Optional metadata to embed
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise ImportError("openpyxl is required for Excel export")

        path = self._prepare_output_path(output_path)
        merged_metadata = self._merge_metadata(metadata)

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        for sheet_name, data in sheets.items():
            if not data:
                continue

            ws = wb.create_sheet(sheet_name)

            # Write headers
            headers = list(data[0].keys())
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

            # Write data
            for row_idx, row_data in enumerate(data, start=2):
                for col_idx, header in enumerate(headers, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Add metadata sheet
        meta_ws = wb.create_sheet('Metadata')
        meta_ws.append(['Key', 'Value'])
        for key, value in merged_metadata.items():
            meta_ws.append([key, str(value)])

        wb.save(path)


class LaTeXExporter(BaseExporter):
    """Export data to LaTeX table format."""

    def export(self, data: List[Dict[str, Any]], output_path: Union[str, Path],
               metadata: Optional[Dict[str, Any]] = None,
               caption: str = '', label: str = '',
               precision: int = 3, max_rows: Optional[int] = None,
               **kwargs) -> None:
        """
        Export data to LaTeX table.

        Args:
            data: List of dictionaries to export
            output_path: Output file path
            metadata: Optional metadata to embed in comments
            caption: Table caption
            label: Table label for referencing
            precision: Decimal precision for floats
            max_rows: Maximum rows to export (None = all)
            **kwargs: Additional options
        """
        if not data:
            raise ValueError("No data to export")

        path = self._prepare_output_path(output_path)
        merged_metadata = self._merge_metadata(metadata)

        # Limit rows if specified
        export_data = data[:max_rows] if max_rows else data

        headers = list(export_data[0].keys())
        num_cols = len(headers)

        with open(path, 'w', encoding='utf-8') as f:
            # Write metadata as comments
            f.write("% Generated LaTeX table\n")
            for key, value in merged_metadata.items():
                f.write(f"% {key}: {value}\n")
            f.write(f"% Total rows: {len(export_data)}\n\n")

            # Begin table
            f.write(r"\begin{table}[htbp]")
            f.write("\n")
            f.write(r"\centering")
            f.write("\n")
            if caption:
                f.write(r"\caption{")
                f.write(caption)
                f.write("}\n")
            if label:
                f.write(r"\label{")
                f.write(label)
                f.write("}\n")

            # Begin tabular
            col_spec = 'l' * num_cols
            f.write(r"\begin{tabular}{")
            f.write(col_spec)
            f.write("}\n")
            f.write(r"\toprule")
            f.write("\n")

            # Write headers
            header_line = ' & '.join([sanitize_latex(str(h)) for h in headers])
            f.write(header_line)
            f.write(r" \\")
            f.write("\n")
            f.write(r"\midrule")
            f.write("\n")

            # Write data rows
            for row_data in export_data:
                row_values = []
                for header in headers:
                    value = row_data.get(header, '')
                    if isinstance(value, float):
                        formatted = f"{value:.{precision}f}"
                    elif isinstance(value, int):
                        formatted = format_number(value)
                    else:
                        formatted = str(value)
                    row_values.append(sanitize_latex(formatted))

                row_line = ' & '.join(row_values)
                f.write(row_line)
                f.write(r" \\")
                f.write("\n")

            # End tabular
            f.write(r"\bottomrule")
            f.write("\n")
            f.write(r"\end{tabular}")
            f.write("\n")
            f.write(r"\end{table}")
            f.write("\n")
